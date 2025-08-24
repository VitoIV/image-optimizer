import re, requests
from typing import List, Tuple, Dict
from PIL import Image
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl

def resize_and_pad(img: Image.Image, min_side: int = 600) -> Image.Image:
    w, h = img.size
    W = max(w, min_side)
    H = max(h, min_side)
    canvas = Image.new("RGB", (W, H), "white")
    x = (W - w) // 2
    y = (H - h) // 2
    canvas.paste(img, (x, y))
    return canvas

def fetch_image(url: str, timeout: int = 25) -> Image.Image:
    r = requests.get(url, stream=True, timeout=timeout)
    r.raise_for_status()
    img = Image.open(BytesIO(r.content))
    if img.mode in ("RGBA","LA") or (img.mode=="P" and "transparency" in img.info):
        img = img.convert("RGB")
    return img

def process_urls(batch_id: str, urls: List[Tuple[int,int,str]], storage, public_base_url: str, threads: int = 8, progress_cb=None) -> Dict[Tuple[int,int], str]:
    results = {}
    total = len(urls)
    done = 0

    def work(item):
        r, c, url = item
        try:
            img = fetch_image(url)
            img = resize_and_pad(img)
            base = url.split("?")[0].rstrip("/").split("/")[-1] or "img"
            base = re.sub(r"[^a-zA-Z0-9]+", "-", base).strip("-").lower() or "img"
            nice_id, out_file = storage.new_image_path(batch_id, base)
            img.save(out_file, format="JPEG", quality=90, optimize=True)
            return (r, c, f"{public_base_url}/api/i/{batch_id}/{nice_id}", None)
        except Exception as e:
            return (r, c, None, str(e))

    with ThreadPoolExecutor(max_workers=max(1, threads)) as ex:
        for fut in as_completed({ex.submit(work, u): u for u in urls}):
            r, c, new_url, err = fut.result()
            if new_url:
                results[(r,c)] = new_url
            done += 1
            if progress_cb:
                progress_cb(done, total)
    return results

def extract_urls_mode_A(ws) -> List[Tuple[int,int,str]]:
    out = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        cell = row[0]
        val = cell.value
        if isinstance(val, str) and val.strip().lower().startswith("http"):
            out.append((i, 1, val.strip()))
    return out

def extract_urls_mode_table(ws) -> List[Tuple[int,int,str]]:
    headers = {idx+1:(cell.value or "") for idx, cell in enumerate(ws[1])}
    picture_cols = [c for c, name in headers.items() if isinstance(name, str) and re.match(r"^PICTURE_\d{1,2}$", name.strip(), re.IGNORECASE)]
    out = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        for col in picture_cols:
            val = ws.cell(row=i, column=col).value
            if isinstance(val, str) and val.strip().lower().startswith("http"):
                out.append((i, col, val.strip()))
    return out

def write_results(ws, mapping: Dict[Tuple[int,int], str]):
    for (r, c), new_url in mapping.items():
        ws.cell(row=r, column=c, value=new_url)
